"""
IntelliRank FastAPI backend.

Endpoints:
  GET  /api/health          — liveness check
  POST /api/candidates/load — load + score candidates (cached)
  POST /api/rank            — rank with optional weight override
  GET  /api/export/xlsx     — export last ranking run as XLSX
  GET  /api/export/csv      — export last ranking run as submission CSV
"""

import csv
import io
import json
import time
from pathlib import Path
from typing import Optional

import numpy as np
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

import sys
sys.path.insert(0, str(Path(__file__).parent.parent))

from pipeline.adapter import normalize
from pipeline import scorer as sc
from pipeline import embeddings as emb_module
from pipeline.jd_config import JD, JD_RAW_TEXT

# ---------------------------------------------------------------------------

app = FastAPI(title="IntelliRank", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# ---------------------------------------------------------------------------
# In-memory state (loaded once per server session)
# ---------------------------------------------------------------------------

_state = {
    "candidates": [],          # list of normalized candidate dicts
    "cosine_sims": {},         # candidate_id → cosine similarity
    "last_ranking": [],        # last ranked output
    "loaded": False,
}

CANDIDATES_PATH = Path(__file__).parent.parent / "data" / "raw" / "candidates.jsonl"


# ---------------------------------------------------------------------------
# Models
# ---------------------------------------------------------------------------

class WeightConfig(BaseModel):
    semantic: float = 0.20
    skill: float = 0.35
    career: float = 0.30
    behavioral: float = 0.15


class RankRequest(BaseModel):
    weights: Optional[WeightConfig] = None
    top_k: int = 100


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _load_candidates_if_needed():
    if _state["loaded"]:
        return

    if not CANDIDATES_PATH.exists():
        raise HTTPException(
            status_code=503,
            detail=f"candidates.jsonl not found at {CANDIDATES_PATH}. "
                   "Copy it to data/raw/ or call /api/candidates/load with a path.",
        )

    # Load embeddings cache
    jd_emb = emb_module.load_jd_embedding()
    cand_embs, cand_ids = emb_module.load_embeddings()

    cosine_sims = {}
    if jd_emb is not None and cand_embs is not None:
        sims = emb_module.cosine_similarities(jd_emb, cand_embs)
        cosine_sims = dict(zip(cand_ids, sims.tolist()))

    candidates = []
    with open(CANDIDATES_PATH, encoding="utf-8") as f:
        for line in f:
            try:
                raw = json.loads(line)
                candidates.append(normalize(raw))
            except Exception:
                pass

    _state["candidates"] = candidates
    _state["cosine_sims"] = cosine_sims
    _state["loaded"] = True


def _score_all(weights: dict) -> list[dict]:
    cosine_sims = _state["cosine_sims"]
    results = []

    for cand in _state["candidates"]:
        cid = cand["candidate_id"]
        skill = sc.skill_match_score(cand)
        career = sc.career_fit_score(cand)
        behavioral = sc.behavioral_score(cand)
        semantic = None
        if cid in cosine_sims:
            semantic = sc.semantic_score_from_cosine(cosine_sims[cid])

        final, effective_weights = sc.fuse_scores(
            semantic=semantic,
            skill=skill,
            career=career,
            behavioral=behavioral,
            weights=weights,
        )

        results.append({
            "candidate_id": cid,
            "name": cand.get("name", ""),
            "headline": cand.get("headline", ""),
            "current_title": cand.get("current_title", ""),
            "years_experience": cand.get("years_experience", 0),
            "country": cand.get("country", ""),
            "location": cand.get("location", ""),
            "skill_names": cand.get("skill_names", []),
            "final_score": final,
            "semantic_score": semantic,
            "skill_match_score": skill,
            "career_fit_score": career,
            "behavioral_score": behavioral,
            "open_to_work": cand.get("open_to_work", False),
            "notice_period_days": cand.get("notice_period_days", 90),
            "recruiter_response_rate": cand.get("recruiter_response_rate", 0),
            "reasoning": sc.build_reasoning(cand, {"final_score": final}),
        })

    results.sort(key=lambda x: (-x["final_score"], x["candidate_id"]))
    return results


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@app.get("/api/health")
def health():
    return {
        "status": "ok",
        "candidates_loaded": _state["loaded"],
        "candidate_count": len(_state["candidates"]),
        "has_embeddings": bool(_state["cosine_sims"]),
    }


@app.post("/api/candidates/load")
def load_candidates():
    _state["loaded"] = False  # force reload
    t0 = time.time()
    _load_candidates_if_needed()
    return {
        "status": "loaded",
        "candidate_count": len(_state["candidates"]),
        "has_embeddings": bool(_state["cosine_sims"]),
        "elapsed_s": round(time.time() - t0, 2),
    }


@app.post("/api/rank")
def rank(req: RankRequest):
    _load_candidates_if_needed()

    weights = {
        "semantic": req.weights.semantic if req.weights else 0.20,
        "skill": req.weights.skill if req.weights else 0.35,
        "career": req.weights.career if req.weights else 0.30,
        "behavioral": req.weights.behavioral if req.weights else 0.15,
    }

    t0 = time.time()
    ranked = _score_all(weights)
    _state["last_ranking"] = ranked

    top_k = min(req.top_k, len(ranked))

    return {
        "jd": {
            "title": JD["title"],
            "company": JD["company"],
            "must_have_clusters": list(JD["must_have_skill_clusters"].keys()),
            "experience_range": f"{JD['min_experience_years']}–{JD['max_experience_years']} years",
        },
        "weights_used": weights,
        "total_candidates": len(ranked),
        "elapsed_s": round(time.time() - t0, 2),
        "results": ranked[:top_k],
    }


@app.get("/api/export/csv")
def export_csv():
    if not _state["last_ranking"]:
        raise HTTPException(status_code=404, detail="No ranking run found. Call /api/rank first.")

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["candidate_id", "rank", "score", "reasoning"])
    for rank_num, r in enumerate(_state["last_ranking"][:100], start=1):
        writer.writerow([
            r["candidate_id"],
            rank_num,
            f"{r['final_score']:.4f}",
            r["reasoning"],
        ])

    buf.seek(0)
    return StreamingResponse(
        io.BytesIO(buf.getvalue().encode("utf-8")),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=ranked_candidates.csv"},
    )


@app.get("/api/export/xlsx")
def export_xlsx():
    if not _state["last_ranking"]:
        raise HTTPException(status_code=404, detail="No ranking run found. Call /api/rank first.")

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ranked Candidates"

    headers = [
        "rank", "candidate_id", "name", "current_title", "years_experience",
        "country", "final_score", "skill_match_score", "career_fit_score",
        "behavioral_score", "semantic_score", "open_to_work",
        "notice_period_days", "top_skills", "reasoning"
    ]

    header_fill = PatternFill("solid", fgColor="7C3AED")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for rank_num, r in enumerate(_state["last_ranking"][:100], start=1):
        row = [
            rank_num,
            r["candidate_id"],
            r.get("name", ""),
            r.get("current_title", ""),
            r.get("years_experience", 0),
            r.get("country", ""),
            round(r["final_score"], 4),
            round(r.get("skill_match_score", 0), 2),
            round(r.get("career_fit_score", 0), 2),
            r.get("behavioral_score"),
            r.get("semantic_score"),
            r.get("open_to_work", False),
            r.get("notice_period_days", ""),
            ", ".join(r.get("skill_names", [])[:5]),
            r.get("reasoning", ""),
        ]
        ws.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=ranked_candidates.xlsx"},
    )

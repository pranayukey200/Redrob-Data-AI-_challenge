"""
generate_xlsx.py — Generate ranked candidates XLSX from pre-computed JSON.
Output: output/ranked_candidates.xlsx
"""
import json
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA  = Path("frontend/public/candidates_data.json")
OUT   = Path("output/ranked_candidates.xlsx")

def score_color(score):
    if score >= 65: return "0055CC"
    if score >= 55: return "1E88E5"
    if score >= 48: return "42A5F5"
    return "90CAF9"

def main():
    with open(DATA, encoding="utf-8") as f:
        data = json.load(f)

    candidates = data["candidates"][:100]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ranked Candidates"

    # ── Header ────────────────────────────────────────────────────────────────
    headers = [
        "Rank", "Candidate ID", "Current Title", "Experience (yrs)",
        "Location", "Country", "Open to Work", "Notice (days)",
        "Response Rate", "Final Score", "Skill Match", "Career Fit",
        "Behavioral", "Semantic", "Top Skills (first 5)", "Reasoning"
    ]
    col_widths = [6, 18, 28, 14, 18, 12, 12, 12, 14, 12, 12, 12, 12, 12, 40, 80]

    hdr_fill  = PatternFill("solid", fgColor="0D1B2A")
    hdr_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.row_dimensions[1].height = 32
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = hdr_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # ── Freeze header row ─────────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── Data rows ─────────────────────────────────────────────────────────────
    for rank_num, c in enumerate(candidates, 1):
        row = rank_num + 1
        score = c["final_score"]
        color_hex = score_color(score)

        score_fill = PatternFill("solid", fgColor=color_hex)
        score_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        data_font  = Font(name="Calibri", size=9)
        center     = Alignment(horizontal="center", vertical="top")
        left       = Alignment(horizontal="left", vertical="top", wrap_text=True)

        top_skills = ", ".join(c.get("skill_names", [])[:5])
        row_color  = "F6F8FA" if rank_num % 2 == 0 else "FFFFFF"
        alt_fill   = PatternFill("solid", fgColor=row_color)

        cells_data = [
            (rank_num,                          center, alt_fill, data_font),
            (c["candidate_id"],                 center, alt_fill, Font(name="Courier New", size=8)),
            (c.get("current_title",""),         left,   alt_fill, data_font),
            (round(c.get("years_experience",0),1), center, alt_fill, data_font),
            ((c.get("location","") or c.get("country","")).split(",")[0], left, alt_fill, data_font),
            (c.get("country",""),               center, alt_fill, data_font),
            ("Yes" if c.get("open_to_work") else "No", center, alt_fill, data_font),
            (c.get("notice_period_days",90),    center, alt_fill, data_font),
            (f"{round(c.get('recruiter_response_rate',0)*100)}%", center, alt_fill, data_font),
            (round(score, 2),                   center, score_fill, score_font),
            (round(c.get("skill_match_score",0),2), center, alt_fill, data_font),
            (round(c.get("career_fit_score",0),2),  center, alt_fill, data_font),
            (round(c.get("behavioral_score",0),2),  center, alt_fill, data_font),
            (round(c["semantic_score"],2) if c.get("semantic_score") is not None else "N/A", center, alt_fill, data_font),
            (top_skills,                        left,   alt_fill, Font(name="Calibri", size=8)),
            (c.get("reasoning",""),             left,   alt_fill, Font(name="Calibri", size=8)),
        ]

        ws.row_dimensions[row].height = 48
        for col_idx, (val, align, fill, font) in enumerate(cells_data, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.alignment = align
            cell.fill = fill
            cell.font = font
            cell.border = border

        # Rank badge: gold for #1, silver for #2-3
        rank_cell = ws.cell(row=row, column=1)
        if rank_num == 1:
            rank_cell.fill = PatternFill("solid", fgColor="D4AF37")
            rank_cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        elif rank_num <= 3:
            rank_cell.fill = PatternFill("solid", fgColor="A8A9AD")
            rank_cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)

    OUT.parent.mkdir(exist_ok=True)
    wb.save(OUT)
    print(f"Saved {OUT}  ({OUT.stat().st_size/1024:.0f} KB)")
    print(f"Top candidate: {candidates[0]['current_title']} — score {candidates[0]['final_score']:.2f}")

if __name__ == "__main__":
    main()
